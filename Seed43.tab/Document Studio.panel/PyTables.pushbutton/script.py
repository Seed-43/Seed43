# -*- coding: utf-8 -*-
"""
PyTables - Link Excel / Word documents to Revit views
Part of the Seed43 extension
IronPython 2 / pyRevit / Revit API
"""
import clr

# PresentationFramework / Core / WindowsBase are not pre-loaded by pyRevit
clr.AddReference('PresentationFramework')
clr.AddReference('PresentationCore')
clr.AddReference('WindowsBase')
# System.Windows.Forms (file/folder dialogs) is also not a standard pyRevit ref
clr.AddReference('System.Windows.Forms')

from pyrevit import revit, DB, script
import System
from System.IO import File, Directory, Path
from System.Windows import Window, Visibility, MessageBox, MessageBoxButton, MessageBoxResult
from System.Windows.Controls import Border, TextBlock, StackPanel, CheckBox, ComboBoxItem
from System.Windows.Media import SolidColorBrush, Color, Colors
from System.Windows.Threading import DispatcherTimer
from System.Windows.Forms import OpenFileDialog, FolderBrowserDialog, DialogResult

import wpf

# Snippets._icons is the shared pyRevit/Seed43 icon library.
# make_icon(key, size, color) returns a WPF UIElement (Canvas with Path shapes).
# Same dependency pattern as LayoutSettings.py in pyTransmit.
try:
    from Snippets._icons import make_icon
    _ICONS_OK = True
except Exception:
    _ICONS_OK = False
    def make_icon(key, size=12, color='#FFFFFF'):
        """Fallback: plain TextBlock when Snippets._icons is unavailable."""
        lbl = TextBlock()
        lbl.Text     = key[:1].upper()
        lbl.FontSize = size
        lbl.Foreground = SolidColorBrush(Color.FromRgb(0xFF, 0xFF, 0xFF))
        return lbl

logger = script.get_logger()
doc    = revit.doc
uidoc  = revit.uidoc


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def safe_str(net_string):
    """Convert a .NET string to a Python 2 str via UTF-8 to avoid codec errors."""
    if net_string is None:
        return ''
    return System.Text.Encoding.UTF8.GetString(
        System.Text.Encoding.UTF8.GetBytes(net_string)
    )


def elem_id_int(eid):
    """Revit 2024+ uses .Value; older uses .IntegerValue."""
    return getattr(eid, 'Value', None) or getattr(eid, 'IntegerValue', None)


def color_from_hex(hex_str):
    """'#208A3C' -> System.Windows.Media.Color"""
    hex_str = hex_str.lstrip('#')
    r = int(hex_str[0:2], 16)
    g = int(hex_str[2:4], 16)
    b = int(hex_str[4:6], 16)
    return Color.FromRgb(r, g, b)


def brush(hex_str):
    return SolidColorBrush(color_from_hex(hex_str))


SCRIPT_DIR = safe_str(Path.GetDirectoryName(__file__))

# Status colours matching the prototype
STATUS_BORDER = {
    'new':   '#6B7280',
    'live':  '#208A3C',
    'stale': '#D69E2E',
    'error': '#C53030',
}
STATUS_LABEL = {
    'new':   'New',
    'live':  'Live',
    'stale': 'Stale',
    'error': 'Error',
}

# Import type keys
IMP_TABLE = 'Table'
IMP_IMAGE = 'Image'

# View type keys
VT_LEGEND   = 'Legend View'
VT_DRAFTING = 'Drafting View'
VT_SCHEDULE = 'Schedule View'

# Conflict policy keys
CP_OVERWRITE = 'Overwrite'
CP_KEEP_BOTH = 'Keep both'
CP_ASK       = 'Ask each time'
CP_SKIP      = 'Skip'


# ---------------------------------------------------------------------------
# Data model: one entry per linked document
# ---------------------------------------------------------------------------
class TableEntry(object):
    def __init__(self):
        self.uid          = safe_str(System.Guid.NewGuid().ToString())
        self.file_path    = ''          # absolute path on disk
        self.sheet_name   = ''          # Excel worksheet name (empty for docx)
        self.range_name   = ''          # named range / '<Used Range>' / '<Print Area>'
        self.view_name    = ''          # target Revit view name
        self.view_type    = VT_LEGEND
        self.import_type  = IMP_TABLE   # Table or Image (set in step 2)
        self.dpi          = 300
        self.scale        = '1:1'
        self.auto_sync    = True
        self.conflict     = CP_ASK
        self.status       = 'new'       # new | live | stale | error
        self.revit_view_id = -1         # ElementId.IntegerValue of created view
        self.last_synced  = ''          # ISO date string

    @property
    def filename(self):
        return safe_str(Path.GetFileName(self.file_path)) if self.file_path else ''

    @property
    def ext(self):
        return safe_str(Path.GetExtension(self.file_path)).lower()

    @property
    def is_excel(self):
        return self.ext in ('.xlsx', '.xls', '.csv')

    @property
    def is_word(self):
        return self.ext in ('.docx', '.doc')

    @property
    def chip_label(self):
        if self.is_excel:
            return 'XLSX'
        if self.is_word:
            return 'DOCX'
        return 'FILE'

    @property
    def chip_color(self):
        if self.is_excel:
            return '#1F7A45'
        if self.is_word:
            return '#2A5DBE'
        return '#555555'

    @property
    def meta_text(self):
        parts = [self.view_type]
        r = self.range_name or '<Used Range>'
        parts.append(r)
        parts.append('{} dpi'.format(self.dpi))
        return u' \u00b7 '.join(parts)

    def to_dict(self):
        return {
            'uid':           self.uid,
            'file_path':     self.file_path,
            'sheet_name':    self.sheet_name,
            'range_name':    self.range_name,
            'view_name':     self.view_name,
            'view_type':     self.view_type,
            'import_type':   self.import_type,
            'dpi':           self.dpi,
            'scale':         self.scale,
            'auto_sync':     self.auto_sync,
            'conflict':      self.conflict,
            'status':        self.status,
            'revit_view_id': self.revit_view_id,
            'last_synced':   self.last_synced,
        }

    @classmethod
    def from_dict(cls, d):
        e = cls()
        for k, v in d.items():
            if hasattr(e, k):
                setattr(e, k, v)
        return e


# ---------------------------------------------------------------------------
# ExtensibleStorage (tracks entries inside the .rvt file)
# ---------------------------------------------------------------------------
SCHEMA_GUID = System.Guid('A3F7C2B1-4D56-4E89-B021-9C8E1F234567')
SCHEMA_NAME = 'PyTablesData'
FIELD_NAME  = 'entries_json'


def _get_or_create_schema():
    from Autodesk.Revit.DB.ExtensibleStorage import Schema, SchemaBuilder, AccessLevel
    existing = Schema.Lookup(SCHEMA_GUID)
    if existing:
        return existing
    sb = SchemaBuilder(SCHEMA_GUID)
    sb.SetSchemaName(SCHEMA_NAME)
    sb.SetReadAccessLevel(AccessLevel.Public)
    sb.SetWriteAccessLevel(AccessLevel.Public)
    sb.AddSimpleField(FIELD_NAME, System.String)
    return sb.Finish()


def load_entries_from_model():
    """Read TableEntry list from ExtensibleStorage. Returns []."""
    try:
        from Autodesk.Revit.DB.ExtensibleStorage import Schema
        schema = Schema.Lookup(SCHEMA_GUID)
        if not schema:
            return []
        data_stores = revit.query.get_elements_by_class(
            DB.DataStorage, doc=doc
        )
        for ds in data_stores:
            entity = ds.GetEntity(schema)
            if not entity.IsValid():
                continue
            raw = safe_str(entity.Get[System.String](FIELD_NAME))
            if not raw:
                return []
            import json
            data = json.loads(raw)
            return [TableEntry.from_dict(d) for d in data]
    except Exception as ex:
        logger.warning('PyTables: could not load entries: {}'.format(ex))
    return []


def save_entries_to_model(entries):
    """Write TableEntry list to ExtensibleStorage in a transaction."""
    try:
        from Autodesk.Revit.DB.ExtensibleStorage import Schema, Entity as ESEntity
        import json
        schema = _get_or_create_schema()
        json_str = json.dumps([e.to_dict() for e in entries])

        data_stores = revit.query.get_elements_by_class(
            DB.DataStorage, doc=doc
        )
        target_ds = None
        for ds in data_stores:
            entity = ds.GetEntity(schema)
            if entity.IsValid():
                target_ds = ds
                break

        with revit.Transaction('PyTables: save links'):
            if target_ds is None:
                target_ds = DB.DataStorage.Create(doc)
            ent = ESEntity(schema)
            ent.Set[System.String](FIELD_NAME, json_str)
            target_ds.SetEntity(ent)
    except Exception as ex:
        logger.warning('PyTables: could not save entries: {}'.format(ex))


# ---------------------------------------------------------------------------
# Excel reading (openpyxl via IronPython — ships with some pyRevit builds)
# ---------------------------------------------------------------------------
def _try_import_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        return None


def _get_workbook_info_xml(file_path):
    """
    Openpyxl-free fallback: read sheet names and named ranges directly
    from the xlsx zip using stdlib zipfile + xml.etree.
    Works in any IronPython 2 environment regardless of pyRevit build.
    Returns {'sheets': [str], 'named_ranges': {sheet_name: [str]}} or None.
    """
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        fp = safe_str(file_path)
        with zipfile.ZipFile(fp) as z:
            wb_xml = z.read('xl/workbook.xml')
        root = ET.fromstring(wb_xml)
        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        sheets = [s.get('name') for s in root.findall('.//ns:sheet', ns)]
        named = {}
        for dn in root.findall('.//ns:definedName', ns):
            rng_name = dn.get('name', '')
            ref = dn.text or ''
            # Skip internal Excel names (_xlnm.*, _FilterDatabase, etc.)
            if rng_name.startswith('_'):
                continue
            if '!' in ref:
                sheet_part = ref.split('!')[0].strip("'").strip('"')
            else:
                sheet_part = sheets[0] if sheets else ''
            if sheet_part:
                named.setdefault(sheet_part, []).append(rng_name)
        return {'sheets': sheets, 'named_ranges': named}
    except Exception as ex:
        logger.warning('PyTables: xml workbook info failed: {}'.format(ex))
        return None


def get_workbook_info(file_path):
    """
    Returns {'sheets': [str], 'named_ranges': {sheet_name: [str]}}.
    Tries openpyxl first (richer data), falls back to zipfile XML parser
    which works in any IronPython 2 environment without third-party libs.
    """
    openpyxl = _try_import_openpyxl()
    if openpyxl:
        try:
            fp = safe_str(file_path)
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            sheets = list(wb.sheetnames)
            named = {}
            for name, defn in wb.defined_names.items():
                if name.startswith('_'):
                    continue
                dests = list(defn.destinations)
                for sheet_title, _ in dests:
                    named.setdefault(sheet_title, []).append(name)
            wb.close()
            return {'sheets': sheets, 'named_ranges': named}
        except Exception as ex:
            logger.warning(
                'PyTables: openpyxl workbook info failed, '
                'trying xml fallback: {}'.format(ex)
            )
    return _get_workbook_info_xml(file_path)



def read_range_data(file_path, sheet_name, range_ref):
    """
    Read cell data from an Excel named range or address.
    Returns list of rows, each row a list of dicts:
        {'value': val, 'bold': bool, 'merge_start': bool}
    range_ref: named range string or '<Used Range>'
    """
    openpyxl = _try_import_openpyxl()
    if not openpyxl:
        return []
    try:
        fp = safe_str(file_path)
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb[sheet_name]

        if range_ref and range_ref not in ('<Used Range>', '<Print Area>'):
            # Named range
            if range_ref in wb.defined_names:
                attr = wb.defined_names[range_ref].attr_text
                # parse 'Sheet1!$G$13:$J$21'
                if '!' in attr:
                    _, cell_range = attr.split('!', 1)
                else:
                    cell_range = attr
                cell_range = cell_range.replace('$', '')
            else:
                cell_range = range_ref
            cells = ws[cell_range]
        else:
            # Used range: iterate all used cells
            cells = list(ws.iter_rows(
                min_row=ws.min_row, max_row=ws.max_row,
                min_col=ws.min_column, max_col=ws.max_column
            ))

        # Collect merged cell ranges for merge detection
        merged = set()
        for mr in ws.merged_cells.ranges:
            for row in mr.cells:
                for cell in row:
                    merged.add(cell.coordinate)

        rows_out = []
        for row in cells:
            row_out = []
            for cell in row:
                row_out.append({
                    'value':       cell.value,
                    'bold':        bool(cell.font and cell.font.bold),
                    'merge_start': cell.coordinate not in merged or cell.value is not None,
                    'col':         cell.column,
                    'row':         cell.row,
                })
            rows_out.append(row_out)
        wb.close()
        return rows_out
    except Exception as ex:
        logger.warning('PyTables: read_range_data failed: {}'.format(ex))
        return []


# ---------------------------------------------------------------------------
# Revit view creation
# ---------------------------------------------------------------------------
def _get_legend_view_type():
    for vft in revit.query.get_elements_by_class(DB.ViewFamilyType, doc=doc):
        if vft.ViewFamily == DB.ViewFamily.Legend:
            return vft.Id
    return None


def _get_drafting_view_type():
    for vft in revit.query.get_elements_by_class(DB.ViewFamilyType, doc=doc):
        if vft.ViewFamily == DB.ViewFamily.Drafting:
            return vft.Id
    return None


def _get_default_text_note_type():
    types = revit.query.get_elements_by_class(DB.TextNoteType, doc=doc)
    types = list(types)
    return types[0].Id if types else None


def create_legend_view(view_name):
    vft_id = _get_legend_view_type()
    if not vft_id:
        raise RuntimeError('No Legend ViewFamilyType found in project')
    with revit.Transaction('PyTables: create legend view'):
        view = doc.CreateLegend(vft_id)
        view.Name = view_name
    return elem_id_int(view.Id)


def create_drafting_view(view_name):
    vft_id = _get_drafting_view_type()
    if not vft_id:
        raise RuntimeError('No Drafting ViewFamilyType found in project')
    with revit.Transaction('PyTables: create drafting view'):
        view = DB.ViewDrafting.Create(doc, vft_id)
        view.Name = view_name
    return elem_id_int(view.Id)


def create_schedule_view(view_name):
    """Create an empty schedule used as a header-only display."""
    cat_id = DB.ElementId(DB.BuiltInCategory.OST_Walls)
    with revit.Transaction('PyTables: create schedule view'):
        view = DB.ViewSchedule.CreateSchedule(doc, cat_id)
        view.Name = view_name
    return elem_id_int(view.Id)


def populate_view_with_table(view_id_int, rows_data, black_white=False):
    """
    Write cell data from rows_data into a Legend or Drafting view
    as TextNote elements arranged in a grid.
    rows_data: list of rows from read_range_data()
    """
    view_id = DB.ElementId(view_id_int)
    view = doc.GetElement(view_id)
    if not view:
        raise RuntimeError('View not found: {}'.format(view_id_int))

    tn_type_id = _get_default_text_note_type()
    if not tn_type_id:
        raise RuntimeError('No TextNoteType found in project')

    # Cell size in Revit internal units (decimal feet)
    cell_w = mm_to_feet(40.0)
    cell_h = mm_to_feet(7.0)
    origin_x = 0.0
    origin_y = 0.0

    with revit.Transaction('PyTables: populate view'):
        for r_idx, row in enumerate(rows_data):
            for c_idx, cell in enumerate(row):
                val = cell.get('value')
                if val is None:
                    continue
                x = origin_x + c_idx * cell_w
                y = origin_y - r_idx * cell_h
                pt = DB.XYZ(x, y, 0)
                opts = DB.TextNoteOptions(tn_type_id)
                opts.HorizontalAlignment = (
                    System.Enum.Parse(
                        System.Type.GetType(
                            'Autodesk.Revit.DB.HorizontalTextAlignment, RevitAPI'
                        ),
                        'Left'
                    )
                )
                text = unicode(val) if val is not None else u''
                DB.TextNote.Create(doc, view_id, pt, text, opts)


def mm_to_feet(mm):
    return mm / 304.8


def populate_schedule_header(view_id_int, rows_data):
    """
    Populate a schedule's header section with row data using
    the same approach as pyTransmit's schedule header population.
    Each row of data becomes a text row in the schedule header.
    """
    view_id = DB.ElementId(view_id_int)
    view = doc.GetElement(view_id)
    if not view:
        raise RuntimeError('View not found: {}'.format(view_id_int))

    with revit.Transaction('PyTables: populate schedule header'):
        tbl_data = view.GetTableData()
        header_data = tbl_data.GetSectionData(
            System.Enum.Parse(
                System.Type.GetType(
                    'Autodesk.Revit.DB.SectionType, RevitAPI'
                ),
                'Header'
            )
        )
        num_rows = len(rows_data)
        num_cols = max(len(r) for r in rows_data) if rows_data else 0
        while header_data.NumberOfRows < num_rows:
            header_data.InsertRow(header_data.NumberOfRows)
        while header_data.NumberOfColumns < num_cols:
            header_data.InsertColumn(header_data.NumberOfColumns)
        for r_idx, row in enumerate(rows_data):
            for c_idx, cell in enumerate(row):
                val = cell.get('value')
                if val is not None:
                    header_data.SetCellText(r_idx, c_idx, unicode(val))


def insert_image_into_view(view_id_int, image_path, entry):
    """
    Insert an image file into a Legend or Drafting view.
    Uses ImageType.Create (available Revit 2020+).
    """
    view_id = DB.ElementId(view_id_int)
    view = doc.GetElement(view_id)
    if not view:
        raise RuntimeError('View not found: {}'.format(view_id_int))

    img_path = safe_str(image_path)
    with revit.Transaction('PyTables: insert image'):
        opts = DB.ImageTypeOptions(img_path, False, DB.ImageTypeSource.Import)
        img_type = DB.ImageType.Create(doc, opts)
        place_opts = DB.ImagePlacementOptions()
        place_opts.PlacementPoint = (
            System.Enum.Parse(
                System.Type.GetType(
                    'Autodesk.Revit.DB.BoxPlacement, RevitAPI'
                ),
                'TopLeft'
            )
        )
        DB.ImageInstance.Create(doc, view, img_type.Id, place_opts)


def render_range_to_image(file_path, sheet_name, range_name, dpi, bw=False):
    """
    Render an Excel range to a PNG using System.Drawing (no Pillow needed).
    Returns the temp PNG path, or None on failure.
    """
    try:
        clr.AddReference('System.Drawing')
        from System.Drawing import (
            Bitmap, Graphics, Font, FontStyle, Brush, Brushes,
            Pen, Color as DColor, SolidBrush, RectangleF, StringFormat,
            StringAlignment, Image as DImage
        )
        from System.Drawing.Imaging import ImageFormat

        rows_data = read_range_data(file_path, sheet_name, range_name)
        if not rows_data:
            return None

        cell_px_w = int(dpi * 1.5)   # ~1.5 inch per cell
        cell_px_h = int(dpi * 0.25)  # ~0.25 inch row height
        num_cols  = max(len(r) for r in rows_data)
        num_rows  = len(rows_data)
        img_w     = cell_px_w * num_cols
        img_h     = cell_px_h * num_rows

        bmp = Bitmap(img_w, img_h)
        g   = Graphics.FromImage(bmp)
        bg  = DColor.White if bw else DColor.FromArgb(43, 51, 64)
        g.Clear(bg)

        font_size = max(8, int(dpi * 0.09))
        font_normal = Font('Segoe UI', font_size, FontStyle.Regular)
        font_bold   = Font('Segoe UI', font_size, FontStyle.Bold)
        fg_color    = DColor.Black if bw else DColor.FromArgb(244, 250, 255)
        fg_brush    = SolidBrush(fg_color)
        fmt = StringFormat()
        fmt.Alignment = StringAlignment.Near
        fmt.LineAlignment = StringAlignment.Center

        for r_idx, row in enumerate(rows_data):
            for c_idx, cell in enumerate(row):
                x = c_idx * cell_px_w
                y = r_idx * cell_px_h
                rect = RectangleF(x + 4, y, cell_px_w - 8, cell_px_h)
                val = cell.get('value')
                if val is None:
                    continue
                text = unicode(val)
                fnt  = font_bold if cell.get('bold') else font_normal
                g.DrawString(text, fnt, fg_brush, rect, fmt)

        g.Dispose()
        tmp_path = safe_str(Path.Combine(
            Path.GetTempPath(),
            'pytables_{}_{}.png'.format(sheet_name, range_name)
        ))
        bmp.Save(tmp_path, ImageFormat.Png)
        bmp.Dispose()
        return tmp_path
    except Exception as ex:
        logger.warning('PyTables: render_range_to_image failed: {}'.format(ex))
        return None


# ---------------------------------------------------------------------------
# Apply entry to Revit
# ---------------------------------------------------------------------------
def apply_entry(entry):
    """
    Create or update the Revit view for a single TableEntry.
    Sets entry.status and entry.revit_view_id on success.
    Returns (True, '') or (False, error_message).
    """
    try:
        file_path = safe_str(entry.file_path)
        if not File.Exists(file_path):
            return False, 'Source file not found: {}'.format(file_path)

        view_name = entry.view_name or safe_str(
            Path.GetFileNameWithoutExtension(file_path)
        )

        # --- Step 1: create the view if it doesn't already exist ---
        if entry.revit_view_id < 0:
            if entry.view_type == VT_LEGEND:
                entry.revit_view_id = create_legend_view(view_name)
            elif entry.view_type == VT_DRAFTING:
                entry.revit_view_id = create_drafting_view(view_name)
            elif entry.view_type == VT_SCHEDULE:
                entry.revit_view_id = create_schedule_view(view_name)

        # --- Step 2: populate content based on import_type ---
        if entry.import_type == IMP_IMAGE:
            img_path = render_range_to_image(
                file_path,
                entry.sheet_name or (
                    get_workbook_info(file_path) or {'sheets': ['Sheet1']}
                )['sheets'][0],
                entry.range_name,
                entry.dpi
            )
            if img_path:
                insert_image_into_view(entry.revit_view_id, img_path)
                try:
                    File.Delete(img_path)
                except Exception:
                    pass
        else:
            # Table mode
            sheet = entry.sheet_name
            if not sheet and entry.is_excel:
                info = get_workbook_info(file_path)
                sheet = info['sheets'][0] if info else 'Sheet1'

            rows_data = read_range_data(file_path, sheet, entry.range_name)
            if entry.view_type == VT_SCHEDULE:
                populate_schedule_header(entry.revit_view_id, rows_data)
            else:
                populate_view_with_table(entry.revit_view_id, rows_data)

        import datetime
        entry.last_synced = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        entry.status = 'live'
        return True, ''

    except Exception as ex:
        entry.status = 'error'
        return False, unicode(ex)


# ---------------------------------------------------------------------------
# Row card builder
# ---------------------------------------------------------------------------
class RowCardController(object):
    """
    Builds and owns the WPF Border that represents one TableEntry row.
    Expand/collapse is toggled manually via Python (no XAML triggers).
    """

    def __init__(self, entry, window_ctrl):
        self.entry   = entry
        self.win     = window_ctrl
        self.is_open = False
        self._wb_info   = None
        self._build()

    def _build(self):
        # ── Outer card ───────────────────────────────────────────────
        self.card = Border()
        self.card.Background      = brush('#323A48')
        self.card.CornerRadius    = System.Windows.CornerRadius(8)
        self.card.Margin          = System.Windows.Thickness(0, 0, 0, 8)
        self.card.BorderThickness = System.Windows.Thickness(3, 0, 0, 0)
        self.card.BorderBrush     = brush(STATUS_BORDER.get(self.entry.status, '#6B7280'))

        outer = StackPanel()
        self.card.Child = outer

        # ── Summary row ──────────────────────────────────────────────
        # Layout: [checkbox] [chip] [name+meta *] [dot] [badge Auto] [caret Auto]
        summary = Border()
        summary.Padding = System.Windows.Thickness(12, 8, 12, 8)
        summary.Cursor  = System.Windows.Input.Cursors.Hand
        summary.MouseLeftButtonUp += self._on_header_click

        SW = System.Windows
        SWC = System.Windows.Controls
        SWS = System.Windows.Shapes

        row_grid = SWC.Grid()

        def _auto_col():
            c = SWC.ColumnDefinition()
            c.Width = SW.GridLength.Auto
            return c

        def _star_col():
            c = SWC.ColumnDefinition()
            c.Width = SW.GridLength(1, SW.GridUnitType.Star)
            return c

        def _px_col(px):
            c = SWC.ColumnDefinition()
            c.Width = SW.GridLength(px, SW.GridUnitType.Pixel)
            return c

        # col 0: checkbox  col 1: spacer  col 2: chip  col 3: spacer
        # col 4: name+meta (star)  col 5: spacer  col 6: dot  col 7: badge  col 8: caret
        for col in [_auto_col(), _px_col(6), _auto_col(), _px_col(8),
                    _star_col(), _px_col(6), _auto_col(), _auto_col(), _auto_col()]:
            row_grid.ColumnDefinitions.Add(col)

        def _set_col(ctrl, idx):
            SWC.Grid.SetColumn(ctrl, idx)
            row_grid.Children.Add(ctrl)

        # Checkbox
        self.check = SWC.CheckBox()
        self.check.VerticalAlignment = SW.VerticalAlignment.Center
        self.check.Checked   += self._on_check_changed
        self.check.Unchecked += self._on_check_changed
        _set_col(self.check, 0)

        # File type chip
        chip = Border()
        chip.Background   = brush(self.entry.chip_color)
        chip.CornerRadius = SW.CornerRadius(4)
        chip.Width  = 36
        chip.Height = 20
        chip.VerticalAlignment = SW.VerticalAlignment.Center
        chip_lbl = TextBlock()
        chip_lbl.Text = self.entry.chip_label
        chip_lbl.FontSize   = 9
        chip_lbl.FontWeight = SW.FontWeights.Bold
        chip_lbl.Foreground = brush('#F4FAFF')
        chip_lbl.HorizontalAlignment = SW.HorizontalAlignment.Center
        chip_lbl.VerticalAlignment   = SW.VerticalAlignment.Center
        chip.Child = chip_lbl
        _set_col(chip, 2)

        # Name + meta (star column — stretches)
        name_stack = SWC.StackPanel()
        name_stack.VerticalAlignment = SW.VerticalAlignment.Center
        self.name_lbl = TextBlock()
        self.name_lbl.Text = self.entry.view_name or safe_str(
            Path.GetFileNameWithoutExtension(self.entry.file_path))
        self.name_lbl.FontSize     = 12.5
        self.name_lbl.FontWeight   = SW.FontWeights.Medium
        self.name_lbl.Foreground   = brush('#F4FAFF')
        self.name_lbl.TextTrimming = SW.TextTrimming.CharacterEllipsis
        self.meta_lbl = TextBlock()
        self.meta_lbl.Text       = self.entry.meta_text
        self.meta_lbl.FontSize   = 10.5
        self.meta_lbl.Foreground = brush('#F4FAFF')
        self.meta_lbl.Opacity    = 0.55
        self.meta_lbl.FontFamily = System.Windows.Media.FontFamily('Consolas')
        self.meta_lbl.Margin     = SW.Thickness(0, 2, 0, 0)
        name_stack.Children.Add(self.name_lbl)
        name_stack.Children.Add(self.meta_lbl)
        _set_col(name_stack, 4)

        # Status dot
        self.dot = SWS.Ellipse()
        self.dot.Width  = 8
        self.dot.Height = 8
        self.dot.Fill   = brush(STATUS_BORDER.get(self.entry.status, '#6B7280'))
        self.dot.VerticalAlignment   = SW.VerticalAlignment.Center
        self.dot.HorizontalAlignment = SW.HorizontalAlignment.Center
        _set_col(self.dot, 6)

        # Status badge
        badge_border = Border()
        badge_border.Background      = brush('#232933')
        badge_border.BorderBrush     = brush('#404553')
        badge_border.BorderThickness = SW.Thickness(1)
        badge_border.CornerRadius    = SW.CornerRadius(999)
        badge_border.Padding         = SW.Thickness(7, 3, 7, 3)
        badge_border.VerticalAlignment = SW.VerticalAlignment.Center
        badge_border.Margin = SW.Thickness(6, 0, 0, 0)
        self.badge_lbl = TextBlock()
        self.badge_lbl.Text      = STATUS_LABEL.get(self.entry.status, 'New')
        self.badge_lbl.FontSize  = 10.5
        self.badge_lbl.Foreground = brush('#F4FAFF')
        self.badge_lbl.Opacity    = 0.85
        badge_border.Child = self.badge_lbl
        _set_col(badge_border, 7)

        # Caret
        self.caret_lbl = TextBlock()
        self.caret_lbl.Text = u'\u25bc'
        self.caret_lbl.FontSize  = 9
        self.caret_lbl.Foreground = brush('#F4FAFF')
        self.caret_lbl.Opacity    = 0.6
        self.caret_lbl.VerticalAlignment   = SW.VerticalAlignment.Center
        self.caret_lbl.HorizontalAlignment = SW.HorizontalAlignment.Center
        self.caret_lbl.Margin = SW.Thickness(8, 0, 0, 0)
        _set_col(self.caret_lbl, 8)

        summary.Child = row_grid
        outer.Children.Add(summary)

        # ── Separator ────────────────────────────────────────────────
        self.sep = System.Windows.Shapes.Rectangle()
        self.sep.Height = 1
        self.sep.Fill = brush('#28303D')
        self.sep.Visibility = Visibility.Collapsed
        outer.Children.Add(self.sep)

        # ── Detail section ───────────────────────────────────────────
        self.detail = Border()
        self.detail.Padding    = System.Windows.Thickness(12, 10, 12, 12)
        self.detail.Visibility = Visibility.Collapsed
        try:
            self._build_detail(self.detail)
        except Exception as ex:
            logger.warning(
                'PyTables: detail build failed for {}: {}'.format(
                    self.entry.filename, ex
                )
            )
            # Provide minimal stubs so update_status does not crash
            self.hint_stale = System.Windows.Controls.TextBlock()
            self.hint_error = System.Windows.Controls.TextBlock()
            self.autosync_track = Border()
            self.autosync_thumb = System.Windows.Shapes.Ellipse()
        outer.Children.Add(self.detail)

    def _build_detail(self, parent):
        """Build the expanded detail form inside the row card."""
        outer = StackPanel()
        parent.Child = outer

        # ── Row 0: View name / View type ─────────────────────────────
        g1 = self._two_col_grid()
        outer.Children.Add(g1)

        self.vname_tb = self._field_box(
            g1, 0, 0, 'View name',
            self.entry.view_name or
            safe_str(Path.GetFileNameWithoutExtension(self.entry.file_path))
        )
        self.vname_tb.TextChanged += self._on_vname_changed

        self.vtype_cb = self._field_combo(
            g1, 0, 1, 'View type',
            [VT_LEGEND, VT_DRAFTING, VT_SCHEDULE],
            self.entry.view_type
        )
        self.vtype_cb.SelectionChanged += self._on_vtype_changed

        # ── Row 1: Region / range / View scale ───────────────────────
        g2 = self._two_col_grid()
        outer.Children.Add(g2)

        self.range_cb = self._combo(['<Used Range>', '<Print Area>'])
        self.range_cb.SelectionChanged += self._on_range_changed
        self._field_add_widget(g2, 0, 0, 'Region / range', self.range_cb)

        self.scale_tb = self._field_box(g2, 0, 1, 'View scale', self.entry.scale)
        self.scale_tb.TextChanged += self._on_scale_changed

        # ── Row 2: DPI / Conflict policy ─────────────────────────────
        g3 = self._two_col_grid()
        outer.Children.Add(g3)

        self.dpi_cb = self._field_combo(
            g3, 0, 0, 'DPI',
            ['150', '200', '300', '400', '600'],
            str(self.entry.dpi)
        )
        self.dpi_cb.SelectionChanged += self._on_dpi_changed

        self.conflict_cb = self._field_combo(
            g3, 0, 1, 'Conflict policy',
            [CP_OVERWRITE, CP_KEEP_BOTH, CP_ASK, CP_SKIP],
            self.entry.conflict
        )
        self.conflict_cb.SelectionChanged += self._on_conflict_changed

        # Sheet selector: hidden, used only for range population
        self.sheet_cb = self._combo([])
        self.sheet_cb.Visibility = System.Windows.Visibility.Collapsed
        self.sheet_cb.SelectionChanged += self._on_sheet_changed
        outer.Children.Add(self.sheet_cb)
        self._populate_sheets()
        # _populate_ranges called inside _populate_sheets when wb_info is available

        # ── Source path (full width) ──────────────────────────────────
        path_section = StackPanel()
        path_section.Margin = System.Windows.Thickness(0, 4, 0, 0)
        path_section.Children.Add(self._label('Source path', '#F4FAFF'))
        self.path_tb = self._textbox(safe_str(self.entry.file_path))
        self.path_tb.IsReadOnly = True
        self.path_tb.FontFamily = System.Windows.Media.FontFamily('Consolas')
        self.path_tb.FontSize   = 10.5
        path_section.Children.Add(self.path_tb)
        outer.Children.Add(path_section)

        # ── Separator ─────────────────────────────────────────────────
        sep = System.Windows.Shapes.Rectangle()
        sep.Height = 1
        sep.Fill   = brush('#28303D')
        sep.Margin = System.Windows.Thickness(0, 12, 0, 10)
        outer.Children.Add(sep)

        # ── Auto-sync + Sync + Remove ─────────────────────────────────
        action_row = System.Windows.Controls.Grid()
        ac1 = System.Windows.Controls.ColumnDefinition()
        ac2 = System.Windows.Controls.ColumnDefinition()
        ac2.Width = System.Windows.GridLength.Auto
        action_row.ColumnDefinitions.Add(ac1)
        action_row.ColumnDefinitions.Add(ac2)

        # Toggle + label
        sync_stack = StackPanel()
        sync_stack.Orientation = System.Windows.Controls.Orientation.Horizontal
        sync_stack.VerticalAlignment = System.Windows.VerticalAlignment.Center

        self.autosync_track = Border()
        self.autosync_track.Width  = 32
        self.autosync_track.Height = 18
        self.autosync_track.CornerRadius    = System.Windows.CornerRadius(999)
        self.autosync_track.BorderThickness = System.Windows.Thickness(1)
        self.autosync_track.Cursor = System.Windows.Input.Cursors.Hand
        self.autosync_track.MouseLeftButtonUp += self._on_autosync_toggle
        self.autosync_thumb = System.Windows.Shapes.Ellipse()
        self.autosync_thumb.Width  = 12
        self.autosync_thumb.Height = 12
        self.autosync_thumb.HorizontalAlignment = (
            System.Windows.HorizontalAlignment.Left
        )
        self.autosync_thumb.Margin = System.Windows.Thickness(2, 0, 0, 0)
        self.autosync_track.Child = self.autosync_thumb
        self._update_autosync_ui()
        sync_stack.Children.Add(self.autosync_track)

        sync_lbl = TextBlock()
        sync_lbl.Text       = 'Auto-sync when file changes'
        sync_lbl.FontSize   = 11.5
        sync_lbl.Foreground = brush('#F4FAFF')
        sync_lbl.Opacity    = 0.85
        sync_lbl.VerticalAlignment = System.Windows.VerticalAlignment.Center
        sync_lbl.Margin = System.Windows.Thickness(8, 0, 0, 0)
        sync_stack.Children.Add(sync_lbl)

        System.Windows.Controls.Grid.SetColumn(sync_stack, 0)
        action_row.Children.Add(sync_stack)

        # Sync + Remove buttons
        btn_stack = StackPanel()
        btn_stack.Orientation = System.Windows.Controls.Orientation.Horizontal
        btn_stack.VerticalAlignment = System.Windows.VerticalAlignment.Center

        sync_btn = self._small_btn(u'\u27f3  Sync', '#208A3C')
        sync_btn.ToolTip = 'Sync now'
        sync_btn.Margin  = System.Windows.Thickness(0, 0, 6, 0)
        sync_btn.Click  += self._on_sync_click
        btn_stack.Children.Add(sync_btn)

        remove_btn = self._small_btn(u'\u2715  Remove', '#C53030')
        remove_btn.ToolTip = 'Remove this link'
        remove_btn.Click  += self._on_remove_click
        btn_stack.Children.Add(remove_btn)

        System.Windows.Controls.Grid.SetColumn(btn_stack, 1)
        action_row.Children.Add(btn_stack)
        outer.Children.Add(action_row)

        # ── Stale / error hints ───────────────────────────────────────
        self.hint_stale = TextBlock()
        self.hint_stale.Text = (
            'Source modified since last sync. '
            'Click Sync to refresh the Revit view.'
        )
        self.hint_stale.FontSize     = 11.5
        self.hint_stale.Foreground   = brush('#D69E2E')
        self.hint_stale.TextWrapping = System.Windows.TextWrapping.Wrap
        self.hint_stale.Margin       = System.Windows.Thickness(0, 10, 0, 0)
        self.hint_stale.Visibility   = (
            Visibility.Visible if self.entry.status == 'stale'
            else Visibility.Collapsed
        )
        outer.Children.Add(self.hint_stale)

        self.hint_error = TextBlock()
        self.hint_error.Text = (
            'View target not found. Re-sync or remove this link.'
        )
        self.hint_error.FontSize     = 11.5
        self.hint_error.Foreground   = brush('#FF8888')
        self.hint_error.TextWrapping = System.Windows.TextWrapping.Wrap
        self.hint_error.Margin       = System.Windows.Thickness(0, 10, 0, 0)
        self.hint_error.Visibility   = (
            Visibility.Visible if self.entry.status == 'error'
            else Visibility.Collapsed
        )
        outer.Children.Add(self.hint_error)

    # ---- UI factory helpers ----
    def _section_label(self, text):
        lbl = TextBlock()
        lbl.Text     = text
        lbl.FontSize = 11
        lbl.FontWeight = System.Windows.FontWeights.SemiBold
        lbl.Foreground = brush('#208A3C')
        lbl.Margin = System.Windows.Thickness(0, 0, 0, 8)
        return lbl

    def _label(self, text, color='#F4FAFF'):
        lbl = TextBlock()
        lbl.Text     = text
        lbl.FontSize = 12
        lbl.FontWeight = System.Windows.FontWeights.Medium
        lbl.Foreground = brush(color)
        lbl.Opacity  = 0.9
        lbl.Margin   = System.Windows.Thickness(0, 0, 0, 4)
        return lbl

    def _textbox(self, text=''):
        tb = System.Windows.Controls.TextBox()
        tb.Text = text
        tb.Background = brush('#F4FAFF')
        tb.Foreground = brush('#2B3340')
        tb.BorderBrush = brush('#208A3C')
        tb.BorderThickness = System.Windows.Thickness(1)
        tb.Padding = System.Windows.Thickness(8, 4, 8, 4)
        tb.FontSize = 12
        tb.Height = 28
        tb.VerticalContentAlignment = System.Windows.VerticalAlignment.Center
        return tb

    def _combo(self, items, selected=None):
        cb = System.Windows.Controls.ComboBox()
        cb.Background = brush('#F4FAFF')
        cb.Foreground = brush('#2B3340')
        cb.BorderBrush = brush('#208A3C')
        cb.BorderThickness = System.Windows.Thickness(1)
        cb.Padding = System.Windows.Thickness(8, 4, 8, 4)
        cb.FontSize = 12
        cb.Height = 28
        cb.Margin = System.Windows.Thickness(0, 0, 0, 0)
        for item in items:
            ci = ComboBoxItem()
            ci.Content = item
            if selected and item == selected:
                ci.IsSelected = True
            cb.Items.Add(ci)
        return cb

    def _small_btn(self, text, bg_color):
        btn = System.Windows.Controls.Button()
        btn.Content  = text
        btn.FontSize = 11
        btn.FontWeight = System.Windows.FontWeights.SemiBold
        btn.Foreground = brush('#F4FAFF')
        btn.Background = brush(bg_color)
        btn.BorderThickness = System.Windows.Thickness(0)
        btn.Padding = System.Windows.Thickness(12, 4, 12, 4)
        btn.Height  = 24
        btn.Cursor  = System.Windows.Input.Cursors.Hand
        btn.Template = self.win._get_rounded_button_template(4)
        return btn

    def _import_radio(self, label, value, parent):
        rb = System.Windows.Controls.RadioButton()
        rb.Content  = label
        rb.Tag      = value
        rb.GroupName = 'import_type_{}'.format(self.entry.uid)
        rb.FontSize = 11
        rb.FontWeight = System.Windows.FontWeights.SemiBold
        rb.Foreground = brush('#F4FAFF')
        rb.Margin = System.Windows.Thickness(0, 0, 6, 0)
        rb.Template = self.win._get_radio_button_template()
        parent.Children.Add(rb)
        return rb

    def _field_add_widget(self, grid, row_idx, col_idx, label_text, widget):
        """Wrap an arbitrary widget with a label and place it in a _two_col_grid."""
        self._ensure_grid_rows(grid, row_idx + 1)
        sp = StackPanel()
        sp.Margin = System.Windows.Thickness(0, 0, 0, 10)
        sp.Children.Add(self._label(label_text))
        sp.Children.Add(widget)
        grid_col = 0 if col_idx == 0 else 2
        System.Windows.Controls.Grid.SetRow(sp, row_idx)
        System.Windows.Controls.Grid.SetColumn(sp, grid_col)
        grid.Children.Add(sp)

    def _two_col_grid(self):
        """Two equal star columns separated by an 8px spacer. Columns 0 and 2 are content."""
        g = System.Windows.Controls.Grid()
        for width, unit in [
            (1,  System.Windows.GridUnitType.Star),
            (8,  System.Windows.GridUnitType.Pixel),
            (1,  System.Windows.GridUnitType.Star),
        ]:
            cd = System.Windows.Controls.ColumnDefinition()
            cd.Width = System.Windows.GridLength(width, unit)
            g.ColumnDefinitions.Add(cd)
        return g

    def _field_box(self, grid, row_idx, col_idx, label_text, value=''):
        """Add label + textbox to a _two_col_grid. col_idx 0 or 1 (maps to grid col 0 or 2)."""
        self._ensure_grid_rows(grid, row_idx + 1)
        sp = StackPanel()
        sp.Margin = System.Windows.Thickness(0, 0, 0, 10)
        sp.Children.Add(self._label(label_text))
        tb = self._textbox(value)
        sp.Children.Add(tb)
        grid_col = 0 if col_idx == 0 else 2
        System.Windows.Controls.Grid.SetRow(sp, row_idx)
        System.Windows.Controls.Grid.SetColumn(sp, grid_col)
        grid.Children.Add(sp)
        return tb

    def _field_combo(self, grid, row_idx, col_idx, label_text, items, selected=None):
        """Add label + combobox to a _two_col_grid. col_idx 0 or 1 (maps to grid col 0 or 2)."""
        self._ensure_grid_rows(grid, row_idx + 1)
        sp = StackPanel()
        sp.Margin = System.Windows.Thickness(0, 0, 0, 10)
        sp.Children.Add(self._label(label_text))
        cb = self._combo(items, selected)
        sp.Children.Add(cb)
        grid_col = 0 if col_idx == 0 else 2
        System.Windows.Controls.Grid.SetRow(sp, row_idx)
        System.Windows.Controls.Grid.SetColumn(sp, grid_col)
        grid.Children.Add(sp)
        return cb

    def _grid_add(self, grid, ctrl, row_idx, col_idx, **kwargs):
        """Place a control in a _two_col_grid. col_idx 0 or 1 (maps to grid col 0 or 2)."""
        self._ensure_grid_rows(grid, row_idx + 1)
        grid_col = 0 if col_idx == 0 else 2
        System.Windows.Controls.Grid.SetRow(ctrl, row_idx)
        System.Windows.Controls.Grid.SetColumn(ctrl, grid_col)
        grid.Children.Add(ctrl)

    @staticmethod
    def _ensure_grid_rows(grid, count):
        while grid.RowDefinitions.Count < count:
            rd = System.Windows.Controls.RowDefinition()
            rd.Height = System.Windows.GridLength.Auto
            grid.RowDefinitions.Add(rd)

    def _populate_sheets(self):
        """Fill the sheet ComboBox with worksheet names from the workbook."""
        self.sheet_cb.Items.Clear()
        if not self.entry.is_excel:
            self.sheet_cb.IsEnabled = False
            self._populate_ranges(None)
            return
        self._wb_info = get_workbook_info(self.entry.file_path)
        info = self._wb_info
        if not info:
            ci = ComboBoxItem()
            ci.Content = 'Sheet1'
            ci.IsSelected = True
            self.sheet_cb.Items.Add(ci)
            self._populate_ranges(None)
            return
        # Ensure entry.sheet_name is set to first sheet if blank
        if not self.entry.sheet_name and info['sheets']:
            self.entry.sheet_name = info['sheets'][0]
        for s in info['sheets']:
            ci = ComboBoxItem()
            ci.Content = s
            if s == self.entry.sheet_name:
                ci.IsSelected = True
            self.sheet_cb.Items.Add(ci)
        # Always call _populate_ranges with full wb_info so named ranges appear
        self._populate_ranges(info)

    def _populate_ranges(self, wb_info=None):
        """Fill the range ComboBox with '<Used Range>' plus any named ranges."""
        self.range_cb.Items.Clear()
        items = ['<Used Range>', '<Print Area>']
        if wb_info:
            nr = wb_info.get('named_ranges', {})
            # Use sheet_name if set, else fall back to all ranges across sheets
            sheet = self.entry.sheet_name
            if sheet and sheet in nr:
                items += nr[sheet]
            elif not sheet:
                seen = set()
                for ranges in nr.values():
                    for r in ranges:
                        if r not in seen:
                            seen.add(r)
                            items.append(r)
        for item in items:
            ci = ComboBoxItem()
            ci.Content = item
            if item == self.entry.range_name or (
                not self.entry.range_name and item == '<Used Range>'
            ):
                ci.IsSelected = True
            self.range_cb.Items.Add(ci)

    # ---- Status update ----
    def update_status(self, status):
        self.entry.status = status
        clr_hex = STATUS_BORDER.get(status, '#6B7280')
        self.card.BorderBrush = brush(clr_hex)
        self.dot.Fill = brush(clr_hex)
        self.badge_lbl.Text = STATUS_LABEL.get(status, status.capitalize())
        self.hint_stale.Visibility = (
            Visibility.Visible if status == 'stale' else Visibility.Collapsed
        )
        self.hint_error.Visibility = (
            Visibility.Visible if status == 'error' else Visibility.Collapsed
        )

    def _update_autosync_ui(self):
        if self.entry.auto_sync:
            self.autosync_track.Background = brush('#3F208A3C')
            self.autosync_track.BorderBrush = brush('#208A3C')
            self.autosync_thumb.Fill = brush('#208A3C')
            self.autosync_thumb.Margin = System.Windows.Thickness(15, 0, 0, 0)
        else:
            self.autosync_track.Background = brush('#1f242c')
            self.autosync_track.BorderBrush = brush('#404553')
            self.autosync_thumb.Fill = brush('#9aa2b1')
            self.autosync_thumb.Margin = System.Windows.Thickness(2, 0, 0, 0)

    # ---- Toggle open/close ----
    def toggle(self):
        self.is_open = not self.is_open
        vis = Visibility.Visible if self.is_open else Visibility.Collapsed
        self.detail.Visibility = vis
        self.sep.Visibility    = vis
        self.caret_lbl.Text = u'\u25b2' if self.is_open else u'\u25bc'

    # ---- Event handlers ----
    def _on_header_click(self, sender, e):
        self.toggle()

    def _on_check_changed(self, sender, e):
        self.win._on_row_check_changed()

    def _on_vname_changed(self, sender, e):
        self.entry.view_name = safe_str(self.vname_tb.Text)
        self.name_lbl.Text   = self.entry.view_name or self.entry.filename

    def _on_vtype_changed(self, sender, e):
        sel = self.vtype_cb.SelectedItem
        if sel:
            self.entry.view_type = safe_str(sel.Content)

    def _on_sheet_changed(self, sender, e):
        sel = self.sheet_cb.SelectedItem
        if sel:
            self.entry.sheet_name = safe_str(sel.Content)
            # Re-use cached wb_info; only re-read if cache is missing
            info = getattr(self, '_wb_info', None) or get_workbook_info(
                self.entry.file_path
            )
            self._populate_ranges(info)

    def _on_range_changed(self, sender, e):
        sel = self.range_cb.SelectedItem
        if sel:
            self.entry.range_name = safe_str(sel.Content)
            self.meta_lbl.Text    = self.entry.meta_text

    def _on_import_type_changed(self, sender, e):
        self.entry.import_type = safe_str(sender.Tag)
        self._rebuild_import_detail()

    def _on_dpi_changed(self, sender, e):
        sel = self.dpi_cb.SelectedItem
        if sel:
            try:
                self.entry.dpi = int(safe_str(sel.Content))
                self.meta_lbl.Text = self.entry.meta_text
            except ValueError:
                pass

    def _on_scale_changed(self, sender, e):
        self.entry.scale = safe_str(self.scale_tb.Text)

    def _on_conflict_changed(self, sender, e):
        sel = self.conflict_cb.SelectedItem
        if sel:
            self.entry.conflict = safe_str(sel.Content)

    def _on_autosync_toggle(self, sender, e):
        self.entry.auto_sync = not self.entry.auto_sync
        self._update_autosync_ui()

    def _on_sync_click(self, sender, e):
        ok, msg = apply_entry(self.entry)
        if ok:
            self.update_status('live')
            save_entries_to_model(self.win.entries)
            self.win.flash('Synced: {}'.format(self.entry.view_name))
        else:
            self.update_status('error')
            self.win.flash('Error: {}'.format(msg))
        self.win._update_footer()

    def _on_remove_click(self, sender, e):
        self.win.remove_entry(self.entry.uid)


# ---------------------------------------------------------------------------
# Main window controller
# ---------------------------------------------------------------------------
class PyTablesWindow(Window):

    def __init__(self):
        xaml_path = safe_str(Path.Combine(SCRIPT_DIR, 'PyTables.xaml'))
        wpf.LoadComponent(self, xaml_path)

        self.entries      = load_entries_from_model()
        self.row_ctrls    = []   # list of RowCardController in display order
        self._toast_timer = None
        self._mode        = 'normal'  # normal | settings | conflicts

        # Wire drag-drop
        self.DragOver += self._on_drag_over
        self.Drop     += self._on_drop

        # Initialise settings defaults
        self._settings_autosync = True
        self._settings_bw       = False

        # Build conflict radio options
        self._build_conflict_options()

        # Render initial state
        self._render_all()
        self._update_footer()

    # ---- Template helpers (used by RowCardController) ----
    def _get_rounded_button_template(self, radius=4):
        """Return a ControlTemplate for a rounded rect button via XamlReader."""
        from System.Windows.Markup import XamlReader
        # xmlns:x required for x:Name and TargetName in Triggers
        xaml = (
            '<ControlTemplate'
            ' xmlns="http://schemas.microsoft.com/winfx/2006/xaml/presentation"'
            ' xmlns:x="http://schemas.microsoft.com/winfx/2006/xaml"'
            ' TargetType="Button">'
            '<Border x:Name="Bd"'
            ' Background="{{TemplateBinding Background}}"'
            ' CornerRadius="{radius}"'
            ' Padding="{{TemplateBinding Padding}}">'.format(radius=radius) +
            '<ContentPresenter HorizontalAlignment="Center"'
            ' VerticalAlignment="Center"/>'
            '</Border>'
            '<ControlTemplate.Triggers>'
            '<Trigger Property="IsMouseOver" Value="True">'
            '<Setter TargetName="Bd" Property="Opacity" Value="0.85"/>'
            '</Trigger>'
            '<Trigger Property="IsPressed" Value="True">'
            '<Setter TargetName="Bd" Property="Opacity" Value="0.70"/>'
            '</Trigger>'
            '</ControlTemplate.Triggers>'
            '</ControlTemplate>'
        )
        return XamlReader.Parse(xaml)

    def _get_radio_button_template(self):
        """Return a ControlTemplate for the import-type RadioButton."""
        from System.Windows.Markup import XamlReader
        # xmlns:x required for x:Name and TargetName in Triggers
        xaml = (
            '<ControlTemplate'
            ' xmlns="http://schemas.microsoft.com/winfx/2006/xaml/presentation"'
            ' xmlns:x="http://schemas.microsoft.com/winfx/2006/xaml"'
            ' TargetType="RadioButton">'
            '<Border x:Name="Bd" Background="#404553"'
            ' CornerRadius="4" Padding="12,5">'
            '<ContentPresenter HorizontalAlignment="Center"'
            ' VerticalAlignment="Center"/>'
            '</Border>'
            '<ControlTemplate.Triggers>'
            '<Trigger Property="IsChecked" Value="True">'
            '<Setter TargetName="Bd" Property="Background" Value="#208A3C"/>'
            '</Trigger>'
            '<Trigger Property="IsMouseOver" Value="True">'
            '<Setter TargetName="Bd" Property="Background" Value="#4E5566"/>'
            '</Trigger>'
            '</ControlTemplate.Triggers>'
            '</ControlTemplate>'
        )
        return XamlReader.Parse(xaml)

    # ---- Conflict panel builder ----
    def _build_conflict_options(self):
        panel = self.FindName('conflict_options_panel')
        if not panel:
            return
        options = [
            (CP_OVERWRITE, 'Revit view is replaced with the new source. Keeps things in sync, loses any in-Revit edits.'),
            (CP_KEEP_BOTH, 'Imports the new source as a copy alongside the existing view. Best when you need to compare.'),
            (CP_ASK,       'Pause sync and ask you what to do each time a change is detected.'),
            (CP_SKIP,      'Ignore the change. View stays stale until you sync manually.'),
        ]
        for name, desc in options:
            row = Border()
            row.Padding     = System.Windows.Thickness(0, 8, 0, 8)
            row.BorderBrush = brush('#404553')
            row.BorderThickness = System.Windows.Thickness(0, 0, 0, 1)
            row_g = System.Windows.Controls.Grid()
            c1 = System.Windows.Controls.ColumnDefinition()
            c1.Width = System.Windows.GridLength.Auto
            c2 = System.Windows.Controls.ColumnDefinition()
            row_g.ColumnDefinitions.Add(c1)
            row_g.ColumnDefinitions.Add(c2)

            rb = System.Windows.Controls.RadioButton()
            rb.GroupName = 'conflict_policy'
            rb.Tag       = name
            rb.Margin    = System.Windows.Thickness(0, 2, 10, 0)
            rb.VerticalAlignment = System.Windows.VerticalAlignment.Top
            rb.IsChecked = (name == CP_ASK)
            System.Windows.Controls.Grid.SetColumn(rb, 0)
            row_g.Children.Add(rb)

            txt_sp = StackPanel()
            title = TextBlock()
            title.Text     = name
            title.FontSize = 12.5
            title.FontWeight = System.Windows.FontWeights.SemiBold
            title.Foreground = brush('#F4FAFF')
            title.Margin = System.Windows.Thickness(0, 0, 0, 2)
            d_lbl = TextBlock()
            d_lbl.Text        = desc
            d_lbl.FontSize    = 11.5
            d_lbl.Foreground  = brush('#F4FAFF')
            d_lbl.Opacity     = 0.65
            d_lbl.TextWrapping = System.Windows.TextWrapping.Wrap
            txt_sp.Children.Add(title)
            txt_sp.Children.Add(d_lbl)
            System.Windows.Controls.Grid.SetColumn(txt_sp, 1)
            row_g.Children.Add(txt_sp)
            row.Child = row_g
            panel.Children.Add(row)

    # ---- Mode switching ----
    def _set_mode(self, mode):
        self._mode = mode
        v_normal  = Visibility.Visible if mode == 'normal'    else Visibility.Collapsed
        v_close   = Visibility.Visible if mode != 'normal'    else Visibility.Collapsed
        v_main    = Visibility.Visible if mode == 'normal'    else Visibility.Collapsed
        v_set     = Visibility.Visible if mode == 'settings'  else Visibility.Collapsed
        v_con     = Visibility.Visible if mode == 'conflicts' else Visibility.Collapsed

        self.FindName('hdr_normal_btns').Visibility = v_normal
        self.FindName('hdr_close_btns').Visibility  = v_close
        self.FindName('main_panel').Visibility      = v_main
        self.FindName('settings_panel').Visibility  = v_set
        self.FindName('conflicts_panel').Visibility = v_con

        sub_set = self.FindName('hdr_sub_settings')
        sub_con = self.FindName('hdr_sub_conflicts')
        if sub_set:
            sub_set.Visibility = Visibility.Visible if mode == 'settings'  else Visibility.Collapsed
        if sub_con:
            sub_con.Visibility = Visibility.Visible if mode == 'conflicts' else Visibility.Collapsed

    # ---- Entry management ----
    def add_files(self, file_paths):
        added = 0
        for fp in file_paths:
            fp = safe_str(fp)
            if not File.Exists(fp):
                continue
            e = TableEntry()
            e.file_path  = fp
            e.view_name  = safe_str(Path.GetFileNameWithoutExtension(fp))
            e.status     = 'new'
            # pre-fill sheet from workbook if Excel
            ext = safe_str(Path.GetExtension(fp)).lower()
            if ext in ('.xlsx', '.xls'):
                info = get_workbook_info(fp)
                if info and info['sheets']:
                    e.sheet_name = info['sheets'][0]
            self.entries.insert(0, e)
            added += 1
        if added:
            self._render_all()
            self._update_footer()
            self.flash('Added {} file{}'.format(added, 's' if added != 1 else ''))

    def remove_entry(self, uid):
        self.entries = [e for e in self.entries if e.uid != uid]
        save_entries_to_model(self.entries)
        self._render_all()
        self._update_footer()
        self.flash('Removed')

    # ---- Render ----
    def _render_all(self):
        """Rebuild the rows_panel from current entries list."""
        rows_panel = self.FindName('rows_panel')
        empty_panel = self.FindName('empty_panel')
        list_panel  = self.FindName('list_panel')
        apply_btn   = self.FindName('apply_btn')

        rows_panel.Children.Clear()
        self.row_ctrls = []

        if not self.entries:
            empty_panel.Visibility = Visibility.Visible
            list_panel.Visibility  = Visibility.Collapsed
            if apply_btn:
                apply_btn.IsEnabled = False
            return

        empty_panel.Visibility = Visibility.Collapsed
        list_panel.Visibility  = Visibility.Visible
        if apply_btn:
            apply_btn.IsEnabled = True

        search_tb = self.FindName('search_tb')
        q = safe_str(search_tb.Text).lower().strip() if search_tb else ''
        for entry in self.entries:
            if q and q not in (entry.view_name or '').lower() \
               and q not in entry.filename.lower() \
               and q not in entry.view_type.lower():
                continue
            try:
                ctrl = RowCardController(entry, self)
                self.row_ctrls.append(ctrl)
                rows_panel.Children.Add(ctrl.card)
            except Exception as ex:
                logger.warning(
                    'PyTables: could not build row for {}: {}'.format(
                        entry.filename, ex
                    )
                )

    def _update_footer(self):
        total = len(self.entries)
        live  = sum(1 for e in self.entries if e.status == 'live')
        stale = sum(1 for e in self.entries if e.status == 'stale')
        error = sum(1 for e in self.entries if e.status == 'error')

        def set_run(name, val):
            r = self.FindName(name)
            if r:
                r.Text = str(val)

        set_run('ft_total_n', total)
        set_run('ft_live_n',  live)
        set_run('ft_stale_n', stale)
        set_run('ft_error_n', error)

        hint = self.FindName('ft_hint')
        if hint:
            hint.Text = 'F5 to sync all' if total > 0 else 'Drop a file to begin'

    def _update_batch_strip(self):
        checked = [c for c in self.row_ctrls if c.check.IsChecked == True]
        strip   = self.FindName('batch_strip')
        remove  = self.FindName('remove_btn')
        if strip:
            strip.Visibility = Visibility.Visible if checked else Visibility.Collapsed
        if remove:
            remove.IsEnabled = bool(checked)
        run = self.FindName('batch_count_run')
        if run:
            run.Text = str(len(checked))

    # ---- Toast ----
    def flash(self, message):
        tb = self.FindName('toast_border')
        tl = self.FindName('toast_text')
        if not tb or not tl:
            return
        tl.Text = message
        tb.Visibility = Visibility.Visible

        if self._toast_timer:
            self._toast_timer.Stop()
        timer = DispatcherTimer()
        timer.Interval = System.TimeSpan.FromMilliseconds(2200)

        def hide(s, e):
            tb.Visibility = Visibility.Collapsed
            timer.Stop()

        timer.Tick += hide
        timer.Start()
        self._toast_timer = timer

    # ---- Settings switch helpers ----
    def _set_switch(self, track_name, thumb_name, lbl_name, on, on_text, off_text):
        track = self.FindName(track_name)
        thumb = self.FindName(thumb_name)
        lbl   = self.FindName(lbl_name)
        if not track:
            return
        if on:
            track.Background  = brush('#3F208A3C')
            track.BorderBrush = brush('#208A3C')
            thumb.Fill        = brush('#208A3C')
            thumb.Margin      = System.Windows.Thickness(15, 0, 0, 0)
            if lbl:
                lbl.Text = on_text
        else:
            track.Background  = brush('#1f242c')
            track.BorderBrush = brush('#404553')
            thumb.Fill        = brush('#9aa2b1')
            thumb.Margin      = System.Windows.Thickness(2, 0, 0, 0)
            if lbl:
                lbl.Text = off_text

    # ======================================================
    # XAML event handlers (wired via x:Name + Click= etc.)
    # ======================================================

    def on_apply_click(self, sender, e):
        """Apply all new/stale entries to Revit."""
        self.FindName('popup_menu').Visibility = Visibility.Collapsed
        to_apply = [en for en in self.entries if en.status in ('new', 'stale')]
        if not to_apply:
            self.flash('Nothing to apply')
            return
        ok_count = 0
        for entry in to_apply:
            ok, msg = apply_entry(entry)
            ctrl = next((c for c in self.row_ctrls if c.entry.uid == entry.uid), None)
            if ctrl:
                ctrl.update_status(entry.status)
            if ok:
                ok_count += 1
        save_entries_to_model(self.entries)
        self._update_footer()
        self.flash('Applied {} view{} to Revit'.format(
            ok_count, 's' if ok_count != 1 else ''))

    def on_menu_click(self, sender, e):
        pm = self.FindName('popup_menu')
        pm.Visibility = (
            Visibility.Visible if pm.Visibility == Visibility.Collapsed
            else Visibility.Collapsed
        )

    def on_panel_close_click(self, sender, e):
        self._set_mode('normal')

    def on_browse_click(self, sender, e):
        self.FindName('popup_menu').Visibility = Visibility.Collapsed
        dlg = OpenFileDialog()
        dlg.Title      = 'Select documents to link'
        dlg.Filter     = 'Supported files|*.xlsx;*.xls;*.csv;*.docx;*.doc|Excel|*.xlsx;*.xls;*.csv|Word|*.docx;*.doc'
        dlg.Multiselect = True
        if dlg.ShowDialog() == DialogResult.OK:
            self.add_files(list(dlg.FileNames))

    def on_folder_click(self, sender, e):
        self.FindName('popup_menu').Visibility = Visibility.Collapsed
        dlg = FolderBrowserDialog()
        dlg.Description = 'Select a folder to import all supported files from'
        if dlg.ShowDialog() == DialogResult.OK:
            folder = safe_str(dlg.SelectedPath)
            files  = []
            for ext in ('*.xlsx', '*.xls', '*.csv', '*.docx', '*.doc'):
                files += [
                    safe_str(f)
                    for f in Directory.GetFiles(folder, ext)
                ]
            if files:
                self.add_files(files)
            else:
                self.flash('No supported files found in that folder')

    def on_search_changed(self, sender, e):
        self._render_all()

    def on_batch_remove_click(self, sender, e):
        checked_uids = [c.entry.uid for c in self.row_ctrls if c.check.IsChecked == True]
        if not checked_uids:
            return
        self.entries = [en for en in self.entries if en.uid not in checked_uids]
        save_entries_to_model(self.entries)
        self._render_all()
        self._update_footer()
        self.flash('Removed {} item{}'.format(
            len(checked_uids), 's' if len(checked_uids) != 1 else ''))

    def on_batch_sync_click(self, sender, e):
        checked = [c for c in self.row_ctrls if c.check.IsChecked == True]
        for ctrl in checked:
            ok, msg = apply_entry(ctrl.entry)
            ctrl.update_status(ctrl.entry.status)
        save_entries_to_model(self.entries)
        self._update_footer()
        self.flash('Synced {} item{}'.format(
            len(checked), 's' if len(checked) != 1 else ''))

    def on_batch_autosync_on_click(self, sender, e):
        for ctrl in self.row_ctrls:
            if ctrl.check.IsChecked == True:
                ctrl.entry.auto_sync = True
                ctrl._update_autosync_ui()
        self.flash('Auto-sync on for selected')

    def on_batch_autosync_off_click(self, sender, e):
        for ctrl in self.row_ctrls:
            if ctrl.check.IsChecked == True:
                ctrl.entry.auto_sync = False
                ctrl._update_autosync_ui()
        self.flash('Auto-sync off for selected')

    def on_menu_settings_click(self, sender, e):
        self.FindName('popup_menu').Visibility = Visibility.Collapsed
        self._set_mode('settings')

    def on_menu_conflicts_click(self, sender, e):
        self.FindName('popup_menu').Visibility = Visibility.Collapsed
        self._set_mode('conflicts')

    def on_sync_all_click(self, sender, e):
        self.FindName('popup_menu').Visibility = Visibility.Collapsed
        count = 0
        for entry in self.entries:
            ok, _ = apply_entry(entry)
            ctrl = next((c for c in self.row_ctrls if c.entry.uid == entry.uid), None)
            if ctrl:
                ctrl.update_status(entry.status)
            if ok:
                count += 1
        save_entries_to_model(self.entries)
        self._update_footer()
        self.flash('Synced {} view{}'.format(count, 's' if count != 1 else ''))

    def on_clear_all_click(self, sender, e):
        self.FindName('popup_menu').Visibility = Visibility.Collapsed
        res = MessageBox.Show(
            'Remove all linked documents?',
            'pyTable',
            MessageBoxButton.YesNo
        )
        if res == MessageBoxResult.Yes:
            self.entries = []
            save_entries_to_model(self.entries)
            self._render_all()
            self._update_footer()
            self.flash('Cleared all documents')

    def on_menu_about_click(self, sender, e):
        self.FindName('popup_menu').Visibility = Visibility.Collapsed
        cnt = self.FindName('about_count')
        if cnt:
            cnt.Text = str(len(self.entries))
        self.FindName('overlay').Visibility     = Visibility.Visible
        self.FindName('about_modal').Visibility = Visibility.Visible

    def on_about_close_click(self, sender, e):
        self.FindName('overlay').Visibility     = Visibility.Collapsed
        self.FindName('about_modal').Visibility = Visibility.Collapsed

    def on_settings_autosync_toggle(self, sender, e):
        self._settings_autosync = not self._settings_autosync
        self._set_switch(
            'settings_autosync_track', 'settings_autosync_thumb', 'settings_autosync_lbl',
            self._settings_autosync,
            'On — files added are kept in sync automatically',
            'Off — files added are kept in sync manually'
        )

    def on_settings_bw_toggle(self, sender, e):
        self._settings_bw = not self._settings_bw
        self._set_switch(
            'settings_bw_track', 'settings_bw_thumb', 'settings_bw_lbl',
            self._settings_bw,
            'Render as black and white by default',
            'Render as black and white by default'
        )

    def _on_row_check_changed(self):
        self._update_batch_strip()

    # ---- Drag and drop ----
    def _on_drag_over(self, sender, e):
        if e.Data.GetDataPresent(System.Windows.DataFormats.FileDrop):
            e.Effects = System.Windows.DragDropEffects.Copy
        else:
            e.Effects = System.Windows.DragDropEffects.None
        e.Handled = True

    def _on_drop(self, sender, e):
        if e.Data.GetDataPresent(System.Windows.DataFormats.FileDrop):
            files = list(e.Data.GetData(System.Windows.DataFormats.FileDrop))
            supported = [
                safe_str(f) for f in files
                if safe_str(Path.GetExtension(f)).lower()
                in ('.xlsx', '.xls', '.csv', '.docx', '.doc')
            ]
            if supported:
                self.add_files(supported)
            else:
                self.flash('No supported files in drop')

    # ---- Show ----
    def show(self):
        self.ShowDialog()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
ctrl = PyTablesWindow()
ctrl.show()
